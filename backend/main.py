from __future__ import annotations

import json
import logging
import os
import random
import re
import shutil
import uuid
from datetime import datetime
from pathlib import Path
from threading import Lock
from typing import Any
from urllib.parse import quote

import pandas as pd
import qrcode
from fastapi import BackgroundTasks, FastAPI, File, Form, Request, UploadFile
from fastapi.encoders import jsonable_encoder
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fpdf import FPDF
from PIL import Image, UnidentifiedImageError
from pydantic import BaseModel, Field, ValidationError, field_validator
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine, make_url
from sqlalchemy.exc import IntegrityError

DEFAULT_DATABASE_URL = "mysql+pymysql://root:root@127.0.0.1:8889/qr_generator"
DATABASE_URL = os.getenv("QR_DATABASE_URL", DEFAULT_DATABASE_URL).strip()

LOGGER = logging.getLogger("qr_generator")
if not LOGGER.handlers:
    logging.basicConfig(
        level=os.getenv("QR_GENERATOR_LOG_LEVEL", "INFO").upper(),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

RNG = random.SystemRandom()
LOT_NUMBER_LOCK = Lock()
DATABASE_READY = False
engine: Engine | None = None

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
FRONTEND_DIR = PROJECT_DIR / "frontend"

PREVIEW_DIR = BASE_DIR / "preview"
EXCEL_DIR = BASE_DIR / "excel"
TEMP_DIR = BASE_DIR / "temp"
DRAFT_META_FILENAME = "meta.json"

MAX_QUANTITY = 5000
MAX_DIGITS = 32
MAX_PREFIX_LENGTH = 20
MAX_LOGO_BYTES = 5 * 1024 * 1024
ALLOWED_LOGO_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}

PDF_MIME_TYPE = "application/pdf"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MIME_TYPE = "text/csv"

SIZE_CODE_MAP = {
    "STD": "STD",
    "M": "M",
    "S": "S",
    "SS": "SS",
}

LABEL_BORDER_RADIUS = 2

LAYOUT_CONFIG = {
    "STD": {
        "orientation": "L",
        "label_w": 25,
        "label_h": 30,
        "label_radius": LABEL_BORDER_RADIUS,
        "cols": 8,
        "qr_w": 23,
        "gap_x": 5,
        "gap_y": 5,
    },
    "M": {
        "orientation": "L",
        "label_w": 45,
        "label_h": 20,
        "label_radius": LABEL_BORDER_RADIUS,
        "cols": 5,
        "qr_w": 15,
        "gap_x": 5,
        "gap_y": 3,
    },
    "S": {
        "orientation": "L",
        "label_w": 40,
        "label_h": 15,
        "label_radius": LABEL_BORDER_RADIUS,
        "cols": 6,
        "qr_w": 12,
        "gap_x": 5,
        "gap_y": 3,
    },
    "SS": {
        "orientation": "P",
        "label_w": 30,
        "label_h": 10,
        "label_radius": LABEL_BORDER_RADIUS,
        "cols": 5,
        "qr_w": 8.5,
        "gap_x": 5,
        "gap_y": 3,
    },
}

OPENPYXL_AVAILABLE = True
try:
    import openpyxl  # noqa: F401
except ImportError:
    OPENPYXL_AVAILABLE = False


class ApiError(Exception):
    def __init__(
        self, status_code: int, message: str, *, details: Any | None = None
    ) -> None:
        super().__init__(message)
        self.status_code = status_code
        self.message = message
        self.details = details


class GenerateConfig(BaseModel):
    mode: str = Field(default="Auto Generate")
    size: str = Field(default="STD")
    quantity: int = Field(default=10, ge=1, le=MAX_QUANTITY)
    digit: int = Field(default=8, ge=1, le=MAX_DIGITS)
    prefix: str = Field(default="VER")
    separator: str = Field(default="")

    @field_validator("mode", mode="before")
    @classmethod
    def normalize_mode(cls, value: Any) -> str:
        normalized = str(value or "Auto Generate").strip()
        if normalized not in {"Auto Generate", "Custom Format"}:
            return "Auto Generate"
        return normalized

    @field_validator("size", mode="before")
    @classmethod
    def normalize_size(cls, value: Any) -> str:
        normalized = str(value or "STD").strip().upper()
        if normalized not in LAYOUT_CONFIG:
            return "STD"
        return normalized

    @field_validator("prefix", mode="before")
    @classmethod
    def normalize_prefix(cls, value: Any) -> str:
        normalized = re.sub(r"\s+", "", str(value or "").strip().upper())
        normalized = re.sub(r"[^A-Z0-9]", "", normalized)
        normalized = normalized[:MAX_PREFIX_LENGTH]
        return normalized or "VER"

    @field_validator("separator", mode="before")
    @classmethod
    def normalize_separator(cls, value: Any) -> str:
        normalized = str(value or "").strip()
        if normalized.lower() == "none":
            return ""
        return normalized if normalized in {"-", "/", "_"} else ""


def ensure_directories() -> None:
    for directory in (
        PREVIEW_DIR,
        EXCEL_DIR,
        TEMP_DIR,
        FRONTEND_DIR / "page",
        FRONTEND_DIR / "css",
        FRONTEND_DIR / "js",
        FRONTEND_DIR / "assets",
    ):
        directory.mkdir(parents=True, exist_ok=True)


ensure_directories()


def quote_mysql_identifier(identifier: str) -> str:
    return f"`{identifier.replace('`', '``')}`"


def make_database_engine(database_url: str) -> Engine | None:
    if not database_url:
        return None

    try:
        return create_engine(database_url, pool_pre_ping=True, future=True)
    except Exception as exc:
        LOGGER.warning("Database engine could not be created: %s", exc)
        return None


def create_database_if_needed(database_url: str) -> None:
    try:
        url = make_url(database_url)
    except Exception as exc:
        LOGGER.warning("Database URL is invalid: %s", exc)
        return

    if not url.drivername.startswith("mysql") or not url.database:
        return

    server_engine: Engine | None = None
    try:
        server_engine = create_engine(
            url.set(database=None), pool_pre_ping=True, future=True
        )
        with server_engine.begin() as connection:
            connection.execute(
                text(
                    "CREATE DATABASE IF NOT EXISTS "
                    f"{quote_mysql_identifier(url.database)} "
                    "CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
                )
            )
    except Exception as exc:
        LOGGER.warning("Could not auto-create database %s: %s", url.database, exc)
    finally:
        if server_engine is not None:
            server_engine.dispose()


def history_table_schema_sql(dialect_name: str) -> str:
    if dialect_name.startswith("mysql"):
        return """
            CREATE TABLE IF NOT EXISTS qr_history (
                id INT NOT NULL AUTO_INCREMENT,
                filename VARCHAR(255) NOT NULL,
                export_filename VARCHAR(255) NULL,
                pdf_data LONGBLOB NOT NULL,
                export_data LONGBLOB NULL,
                export_mime_type VARCHAR(100) NULL,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (id),
                UNIQUE KEY uq_qr_history_filename (filename),
                KEY idx_qr_history_export_filename (export_filename)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """

    return """
        CREATE TABLE IF NOT EXISTS qr_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL UNIQUE,
            export_filename TEXT,
            pdf_data BLOB NOT NULL,
            export_data BLOB,
            export_mime_type TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """


def get_existing_history_columns(connection) -> set[str]:
    if not engine:
        return set()

    if engine.dialect.name.startswith("mysql"):
        rows = connection.execute(
            text(
                """
                SELECT COLUMN_NAME
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE()
                  AND TABLE_NAME = 'qr_history'
                """
            )
        )
        return {str(row[0]) for row in rows}

    rows = connection.execute(text("PRAGMA table_info(qr_history)"))
    return {str(row[1]) for row in rows}


def prepare_history_table(connection) -> None:
    required_columns = {
        "id",
        "filename",
        "export_filename",
        "pdf_data",
        "export_data",
        "export_mime_type",
        "created_at",
    }
    existing_columns = get_existing_history_columns(connection)

    if existing_columns and existing_columns != required_columns:
        if engine and engine.dialect.name.startswith("mysql"):
            legacy_name = f"qr_history_legacy_{datetime.now():%Y%m%d%H%M%S}"
            connection.execute(
                text(
                    "RENAME TABLE qr_history TO "
                    f"{quote_mysql_identifier(legacy_name)}"
                )
            )
            LOGGER.warning(
                "Archived old qr_history table as %s before creating blob storage table.",
                legacy_name,
            )
        else:
            connection.execute(text("DROP TABLE qr_history"))

    connection.execute(
        text(history_table_schema_sql(engine.dialect.name if engine else "mysql"))
    )


def init_database() -> bool:
    global DATABASE_READY, engine

    DATABASE_READY = False
    if not DATABASE_URL:
        LOGGER.info("Database disabled because QR_DATABASE_URL is empty.")
        return False

    create_database_if_needed(DATABASE_URL)

    if engine is None:
        engine = make_database_engine(DATABASE_URL)
    if engine is None:
        return False

    try:
        with engine.begin() as connection:
            prepare_history_table(connection)
        DATABASE_READY = True
        LOGGER.info("Database is ready.")
        return True
    except Exception as exc:
        LOGGER.warning("Database is unavailable: %s", exc)
        DATABASE_READY = False
        return False


def ensure_database_ready() -> bool:
    if DATABASE_READY:
        return True
    return init_database()


app = FastAPI(title="QR Generator API", version="2.0.0")


@app.on_event("startup")
def startup_event() -> None:
    init_database()


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/preview", StaticFiles(directory=str(PREVIEW_DIR)), name="preview")
app.mount("/excel", StaticFiles(directory=str(EXCEL_DIR)), name="excel")
app.mount(
    "/page", StaticFiles(directory=str(FRONTEND_DIR / "page"), html=True), name="page"
)
app.mount("/css", StaticFiles(directory=str(FRONTEND_DIR / "css")), name="css")
app.mount("/js", StaticFiles(directory=str(FRONTEND_DIR / "js")), name="js")
app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIR / "assets")), name="assets")


def api_success(
    data: dict[str, Any] | None = None, status_code: int = 200
) -> JSONResponse:
    payload = {"success": True}
    if data:
        payload.update(data)
    return JSONResponse(status_code=status_code, content=jsonable_encoder(payload))


def api_error(
    status_code: int, message: str, *, details: Any | None = None
) -> JSONResponse:
    payload: dict[str, Any] = {"success": False, "error": message}
    if details is not None:
        payload["details"] = details
    return JSONResponse(status_code=status_code, content=jsonable_encoder(payload))


@app.exception_handler(ApiError)
async def api_error_handler(_: Request, exc: ApiError) -> JSONResponse:
    return api_error(exc.status_code, exc.message, details=exc.details)


@app.exception_handler(RequestValidationError)
async def request_validation_handler(
    _: Request, exc: RequestValidationError
) -> JSONResponse:
    return api_error(422, "Invalid request", details=exc.errors())


def build_file_url(request: Request, route_prefix: str, filename: str) -> str:
    return f"{str(request.base_url).rstrip('/')}{route_prefix}/{quote(filename)}"


def history_record_timestamp(value: Any) -> float | None:
    if isinstance(value, datetime):
        return value.timestamp()
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).timestamp()
        except ValueError:
            return None
    return None


def get_database_engine() -> Engine:
    if not ensure_database_ready() or engine is None:
        raise ApiError(503, "Database is unavailable")
    return engine


def get_export_mime_type(filename: str) -> str:
    return CSV_MIME_TYPE if Path(filename).suffix.lower() == ".csv" else XLSX_MIME_TYPE


def content_disposition(disposition: str, filename: str) -> str:
    return f"{disposition}; filename*=UTF-8''{quote(Path(filename).name)}"


def insert_history_record(
    filename: str,
    pdf_data: bytes,
    export_filename: str | None = None,
    export_data: bytes | None = None,
    export_mime_type: str | None = None,
) -> None:
    db_engine = get_database_engine()

    try:
        with db_engine.begin() as connection:
            exists = connection.execute(
                text("SELECT 1 FROM qr_history WHERE filename = :filename LIMIT 1"),
                {"filename": filename},
            ).first()
            if exists:
                raise ApiError(
                    409,
                    "A file with this name already exists. Please generate a new preview.",
                )

            connection.execute(
                text(
                    """
                    INSERT INTO qr_history (
                        filename,
                        export_filename,
                        pdf_data,
                        export_data,
                        export_mime_type
                    )
                    VALUES (
                        :filename,
                        :export_filename,
                        :pdf_data,
                        :export_data,
                        :export_mime_type
                    )
                    """
                ),
                {
                    "filename": filename,
                    "export_filename": export_filename,
                    "pdf_data": pdf_data,
                    "export_data": export_data,
                    "export_mime_type": export_mime_type,
                },
            )
    except ApiError:
        raise
    except IntegrityError as exc:
        raise ApiError(
            409, "A file with this name already exists. Please generate a new preview."
        ) from exc
    except Exception as exc:
        LOGGER.exception("Failed to insert history record into database")
        raise ApiError(500, "Failed to save confirmed files to database") from exc


def delete_history_record(filename: str) -> bool:
    db_engine = get_database_engine()

    try:
        with db_engine.begin() as connection:
            result = connection.execute(
                text("DELETE FROM qr_history WHERE filename = :filename"),
                {"filename": filename},
            )
            return bool(result.rowcount)
    except Exception as exc:
        LOGGER.exception("Failed to delete history record from database")
        raise ApiError(500, "Failed to delete history record from database") from exc


def load_history_records() -> list[dict[str, Any]]:
    db_engine = get_database_engine()

    try:
        with db_engine.begin() as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT filename, export_filename, created_at
                    FROM qr_history
                    ORDER BY created_at DESC, id DESC
                    """
                )
            ).mappings()
            return [dict(row) for row in rows]
    except Exception as exc:
        LOGGER.exception("Failed to load history records from database")
        raise ApiError(500, "Failed to load history records from database") from exc


def load_history_filenames() -> list[str]:
    try:
        db_engine = get_database_engine()
        with db_engine.begin() as connection:
            rows = connection.execute(
                text("SELECT filename FROM qr_history")
            ).fetchall()
            return [str(row[0]) for row in rows]
    except ApiError:
        return []
    except Exception:
        LOGGER.exception("Failed to load history filenames from database")
        return []


def load_pdf_data(filename: str) -> bytes:
    db_engine = get_database_engine()
    with db_engine.begin() as connection:
        row = connection.execute(
            text("SELECT pdf_data FROM qr_history WHERE filename = :filename LIMIT 1"),
            {"filename": filename},
        ).first()

    if not row or row[0] is None:
        raise ApiError(404, "PDF file not found")
    return bytes(row[0])


def load_export_data(filename: str) -> tuple[bytes, str]:
    db_engine = get_database_engine()
    with db_engine.begin() as connection:
        row = connection.execute(
            text(
                """
                SELECT export_data, export_mime_type
                FROM qr_history
                WHERE export_filename = :filename
                LIMIT 1
                """
            ),
            {"filename": filename},
        ).first()

    if not row or row[0] is None:
        raise ApiError(404, "Export file not found")
    return bytes(row[0]), str(row[1] or get_export_mime_type(filename))


def is_valid_draft_id(draft_id: str | None) -> bool:
    return bool(re.fullmatch(r"[a-f0-9]{32}", str(draft_id or "")))


def get_draft_dir(draft_id: str) -> Path:
    if not is_valid_draft_id(draft_id):
        raise ApiError(400, "Invalid draft id")
    return TEMP_DIR / draft_id


def safe_child_path(
    base_dir: Path, filename: str, allowed_extensions: set[str]
) -> Path:
    candidate = (base_dir / Path(filename or "").name).resolve()
    base_path = base_dir.resolve()

    if candidate.parent != base_path:
        raise ApiError(400, "Invalid file path")
    if candidate.suffix.lower() not in allowed_extensions:
        raise ApiError(400, "Invalid file type")
    return candidate


def read_json_file(file_path: Path) -> dict[str, Any]:
    if not file_path.exists():
        raise FileNotFoundError(file_path.name)
    try:
        return json.loads(file_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        LOGGER.exception("Invalid JSON metadata in %s", file_path)
        raise ApiError(500, "Stored metadata is corrupted") from exc


def write_json_file(file_path: Path, payload: dict[str, Any]) -> None:
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8"
    )


def read_draft_metadata(draft_id: str) -> dict[str, Any]:
    draft_dir = get_draft_dir(draft_id)
    meta_path = draft_dir / DRAFT_META_FILENAME
    if not meta_path.exists():
        raise FileNotFoundError("Draft metadata not found")
    return read_json_file(meta_path)


def write_draft_metadata(draft_dir: Path, metadata: dict[str, Any]) -> None:
    write_json_file(draft_dir / DRAFT_META_FILENAME, metadata)


def remove_draft(draft_id: str) -> None:
    draft_dir = get_draft_dir(draft_id)
    if draft_dir.exists():
        shutil.rmtree(draft_dir, ignore_errors=True)


def sanitize_company_code(value: str) -> str:
    sanitized = re.sub(r"[^A-Z0-9]", "", str(value or "").upper())
    return sanitized[:10] or "VER"


def get_mode_code(mode: str) -> str:
    return "C" if str(mode).strip() == "Custom Format" else "A"


def get_qr_size_code(size_type: str) -> str:
    return SIZE_CODE_MAP.get(str(size_type or "STD").upper(), "STD")


def extract_lot_number_from_filename(
    filename: str,
    company_code: str,
    mode_code: str,
    qr_size_code: str,
) -> int | None:
    base_name = Path(filename).stem
    parts = base_name.split("-")
    if len(parts) != 4:
        return None

    file_company, file_lot, file_mode, file_size = parts
    if (
        file_company != company_code
        or file_mode != mode_code
        or file_size != qr_size_code
        or not file_lot.isdigit()
    ):
        return None
    return int(file_lot)


def collect_reserved_lot_numbers(
    company_code: str, mode_code: str, qr_size_code: str
) -> set[int]:
    reserved: set[int] = set()
    for meta_path in TEMP_DIR.glob(f"*/{DRAFT_META_FILENAME}"):
        try:
            metadata = read_json_file(meta_path)
            filename = str(metadata.get("final_pdf_filename") or "")
            lot_number = extract_lot_number_from_filename(
                filename,
                company_code,
                mode_code,
                qr_size_code,
            )
            if lot_number is not None:
                reserved.add(lot_number)
        except Exception:
            LOGGER.exception("Failed to read draft metadata from %s", meta_path)
    return reserved


def get_next_lot_no(company_code: str, mode_code: str, qr_size_code: str) -> str:
    max_lot = 0

    for pdf_path in PREVIEW_DIR.glob("*.pdf"):
        lot_number = extract_lot_number_from_filename(
            pdf_path.name,
            company_code,
            mode_code,
            qr_size_code,
        )
        if lot_number is not None:
            max_lot = max(max_lot, lot_number)

    for filename in load_history_filenames():
        lot_number = extract_lot_number_from_filename(
            filename,
            company_code,
            mode_code,
            qr_size_code,
        )
        if lot_number is not None:
            max_lot = max(max_lot, lot_number)

    for lot_number in collect_reserved_lot_numbers(
        company_code, mode_code, qr_size_code
    ):
        max_lot = max(max_lot, lot_number)

    return str(max_lot + 1).zfill(4)


def generate_codes(config: GenerateConfig) -> list[str]:
    codes: set[str] = set()

    if config.mode == "Auto Generate":
        while len(codes) < config.quantity:
            raw_number = "".join(str(RNG.randint(0, 9)) for _ in range(11))
            codes.add(f"{raw_number[:3]}-{raw_number[3:7]}-{raw_number[7:]}")
    else:
        while len(codes) < config.quantity:
            digits = "".join(str(RNG.randint(0, 9)) for _ in range(config.digit))
            code = (
                f"{config.prefix}{config.separator}{digits}"
                if config.prefix
                else digits
            )
            codes.add(code)

    return sorted(codes)


async def save_logo_file(logo: UploadFile, target_dir: Path) -> Path:
    extension = Path(logo.filename or "").suffix.lower()
    if extension not in ALLOWED_LOGO_EXTENSIONS:
        raise ApiError(400, "Unsupported logo format")

    content = await logo.read()
    if not content:
        raise ApiError(400, "Logo file is empty")
    if len(content) > MAX_LOGO_BYTES:
        raise ApiError(400, "Logo file is too large")

    logo_path = target_dir / f"logo_{uuid.uuid4().hex}{extension}"
    logo_path.write_bytes(content)

    try:
        with Image.open(logo_path) as image:
            image.verify()
    except (UnidentifiedImageError, OSError) as exc:
        logo_path.unlink(missing_ok=True)
        raise ApiError(400, "Logo file is not a valid image") from exc

    return logo_path


def generate_qr_image(
    code: str, qr_dir: Path, logo_path: Path | None = None
) -> tuple[Path, str]:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=2,
    )
    qr.add_data(code)
    qr.make(fit=True)

    qr_image = qr.make_image(fill_color="black", back_color="white").convert("RGBA")

    if logo_path and logo_path.exists():
        with Image.open(logo_path).convert("RGBA") as logo_image:
            qr_width, qr_height = qr_image.size
            max_logo_size = max(24, qr_width // 4)
            logo_image.thumbnail((max_logo_size, max_logo_size), Image.LANCZOS)

            background_size = int(max(logo_image.size) * 1.2)
            background = Image.new(
                "RGBA", (background_size, background_size), (255, 255, 255, 255)
            )
            paste_x = (background_size - logo_image.width) // 2
            paste_y = (background_size - logo_image.height) // 2
            background.paste(logo_image, (paste_x, paste_y), logo_image)

            qr_position = (
                (qr_width - background_size) // 2,
                (qr_height - background_size) // 2,
            )
            qr_image.paste(background, qr_position, background)

    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", code).strip("_")[:80] or "qr"
    unique_name = f"{safe_name}_{uuid.uuid4().hex[:10]}"
    file_path = qr_dir / f"{unique_name}.png"
    qr_image.save(file_path)
    return file_path, unique_name


def try_register_mitr_fonts(pdf: FPDF) -> bool:
    regular_path = BASE_DIR / "Mitr-Regular.ttf"
    bold_path = BASE_DIR / "Mitr-SemiBold.ttf"
    if regular_path.exists() and bold_path.exists():
        pdf.add_font("Mitr", "", str(regular_path), uni=True)
        pdf.add_font("Mitr", "B", str(bold_path), uni=True)
        return True
    return False


def draw_label_outline(
    pdf: FPDF,
    x: float,
    y: float,
    width: float,
    height: float,
    radius: float,
) -> None:
    if radius <= 0:
        pdf.rect(x, y, width, height)
        return

    try:
        pdf.rect(
            x,
            y,
            width,
            height,
            round_corners=True,
            corner_radius=radius,
        )
    except TypeError:
        rounded_rect = getattr(pdf, "rounded_rect", None)
        if callable(rounded_rect):
            rounded_rect(x, y, width, height, radius)
        else:
            pdf.rect(x, y, width, height)


def create_pdf_layout(
    codes: list[dict[str, str]],
    qr_dir: Path,
    size_type: str,
    prefix: str,
    mode: str,
    output_dir: Path,
) -> str:
    config = LAYOUT_CONFIG[size_type]
    orientation = config["orientation"]
    pdf = FPDF(orientation=orientation, unit="mm", format="A4")
    pdf.set_auto_page_break(False)

    has_mitr_font = try_register_mitr_fonts(pdf)
    page_width = 297 if orientation == "L" else 210
    page_height = 210 if orientation == "L" else 297

    margin_x = 10
    margin_top = 10
    margin_bottom = 15

    label_width = config["label_w"]
    label_height = config["label_h"]
    label_radius = config["label_radius"]
    qr_size = config["qr_w"]
    column_count = config["cols"]
    gap_x = config["gap_x"]
    gap_y = config["gap_y"]

    step_x = label_width + gap_x
    step_y = label_height + gap_y
    usable_height = page_height - margin_top - margin_bottom
    row_count = max(1, int((usable_height + gap_y) // step_y))
    per_page = column_count * row_count
    total_pages = max(1, (len(codes) + per_page - 1) // per_page)

    mode_code = get_mode_code(mode)
    qr_size_code = get_qr_size_code(size_type)
    company_code = sanitize_company_code(prefix) if mode_code == "C" else "VER"
    lot_number = get_next_lot_no(company_code, mode_code, qr_size_code)

    grid_width = (column_count * label_width) + ((column_count - 1) * gap_x)
    grid_height = (row_count * label_height) + ((row_count - 1) * gap_y)
    start_x = max(margin_x, (page_width - grid_width) / 2)
    start_y = max(margin_top, (page_height - grid_height) / 2)

    for page_index in range(total_pages):
        pdf.add_page()
        page_codes = codes[page_index * per_page : (page_index + 1) * per_page]

        for index, code_data in enumerate(page_codes):
            column = index % column_count
            row = index // column_count
            x = start_x + (column * step_x)
            y = start_y + (row * step_y)

            pdf.set_draw_color(200, 200, 200)
            draw_label_outline(pdf, x, y, label_width, label_height, label_radius)

            image_path = qr_dir / f"{code_data['safe_name']}.png"
            if not image_path.exists():
                continue

            if has_mitr_font:
                font_size = 6 if size_type == "SS" else 7 if size_type == "S" else 8
                pdf.set_font("Mitr", size=font_size)
            else:
                font_size = 6 if size_type == "SS" else 7 if size_type == "S" else 8
                pdf.set_font("Arial", size=font_size)

            if size_type == "STD":
                pdf.set_xy(x, y + 1.5)
                pdf.cell(label_width, 4, code_data["code"], align="C")
                qr_x = x + (label_width - qr_size) / 2
                qr_y = y + 6
            else:
                qr_x = x + 1
                qr_y = y + (label_height - qr_size) / 2
                text_x = qr_x + qr_size + 1
                pdf.set_xy(text_x, y)
                pdf.cell(
                    label_width - qr_size - 2,
                    label_height,
                    code_data["code"],
                    align="L",
                )

            pdf.image(str(image_path), x=qr_x, y=qr_y, w=qr_size, h=qr_size)

        stock_code = (
            f"{company_code}-{lot_number}-{mode_code}-{qr_size_code}-{page_index + 1}"
        )
        pdf.set_font("Arial", "B", 16)
        pdf.set_xy(margin_x, page_height - 15)
        pdf.cell(page_width - (margin_x * 2), 8, stock_code, align="R")

    filename = f"{company_code}-{lot_number}-{mode_code}-{qr_size_code}.pdf"
    output_path = output_dir / filename
    pdf.output(str(output_path))

    if not output_path.exists():
        raise ApiError(500, "Failed to create PDF")

    LOGGER.info("PDF created successfully: %s", output_path)
    return filename


def export_codes_file(
    codes: list[str], pdf_filename: str, output_dir: Path
) -> tuple[str, str | None]:
    base_name = Path(pdf_filename).stem
    export_warning = None
    dataframe = pd.DataFrame(codes, columns=["Unique_Code"])

    if OPENPYXL_AVAILABLE:
        export_filename = f"{base_name}.xlsx"
        dataframe.to_excel(output_dir / export_filename, index=False)
    else:
        export_filename = f"{base_name}.csv"
        dataframe.to_csv(output_dir / export_filename, index=False)
        export_warning = (
            "openpyxl is not installed, so the export was generated as CSV."
        )

    return export_filename, export_warning


def get_draft_asset_urls(
    request: Request,
    draft_id: str,
    pdf_filename: str,
    export_filename: str | None,
) -> dict[str, str | None]:
    return {
        "pdf_url": build_file_url(request, f"/draft/preview/{draft_id}", pdf_filename),
        "excel_url": (
            build_file_url(request, f"/draft/export/{draft_id}", export_filename)
            if export_filename
            else None
        ),
    }


def run_generate_job(
    draft_id: str, config_data: GenerateConfig, logo_path: Path | None = None
) -> None:
    draft_dir = get_draft_dir(draft_id)
    qr_dir = draft_dir / "qrcodes"
    qr_dir.mkdir(parents=True, exist_ok=True)

    try:
        raw_codes = generate_codes(config_data)
        rendered_codes: list[dict[str, str]] = []

        for code in raw_codes:
            _, safe_name = generate_qr_image(code, qr_dir, logo_path)
            rendered_codes.append({"code": code, "safe_name": safe_name})

        # Lot numbers are derived from existing output. Keep this section serialized so
        # simultaneous preview jobs cannot reserve the same final filename.
        with LOT_NUMBER_LOCK:
            pdf_filename = create_pdf_layout(
                rendered_codes,
                qr_dir,
                config_data.size,
                config_data.prefix,
                config_data.mode,
                draft_dir,
            )
            export_filename, export_warning = export_codes_file(
                raw_codes, pdf_filename, draft_dir
            )

            write_draft_metadata(
                draft_dir,
                {
                    "draft_id": draft_id,
                    "status": "ready",
                    "final_pdf_filename": pdf_filename,
                    "export_filename": export_filename,
                    "warning": export_warning,
                },
            )
    except Exception as exc:
        LOGGER.exception("Generate job failed for draft %s", draft_id)
        write_draft_metadata(
            draft_dir,
            {
                "draft_id": draft_id,
                "status": "error",
                "error": str(exc) or "Failed to generate preview",
            },
        )
    finally:
        if logo_path and logo_path.exists():
            logo_path.unlink(missing_ok=True)


def find_draft_pdf(draft_dir: Path, draft_id: str) -> str | None:
    try:
        metadata = read_draft_metadata(draft_id)
        filename = str(metadata.get("final_pdf_filename") or "")
        if filename:
            return filename
    except FileNotFoundError:
        pass

    matches = sorted(file.name for file in draft_dir.glob("*.pdf") if file.is_file())
    return matches[0] if matches else None


def find_draft_export(draft_dir: Path, draft_id: str) -> str | None:
    try:
        metadata = read_draft_metadata(draft_id)
        filename = str(metadata.get("export_filename") or "")
        if filename:
            return filename
    except FileNotFoundError:
        pass

    matches = sorted(
        file.name
        for file in draft_dir.iterdir()
        if file.is_file() and file.suffix.lower() in {".xlsx", ".csv"}
    )
    return matches[0] if matches else None


@app.post("/generate")
async def generate(
    request: Request,
    background_tasks: BackgroundTasks,
    config: str = Form(...),
    logo: UploadFile | None = File(default=None),
    draft_id: str | None = Form(default=None),
) -> JSONResponse:
    active_draft_id = draft_id if is_valid_draft_id(draft_id) else uuid.uuid4().hex
    draft_dir = get_draft_dir(active_draft_id)
    qr_dir = draft_dir / "qrcodes"
    draft_dir.mkdir(parents=True, exist_ok=True)
    qr_dir.mkdir(parents=True, exist_ok=True)

    logo_path: Path | None = None

    try:
        config_data = GenerateConfig.model_validate_json(config)
        if logo and logo.filename:
            logo_path = await save_logo_file(logo, draft_dir)

        write_draft_metadata(
            draft_dir,
            {
                "draft_id": active_draft_id,
                "status": "processing",
            },
        )

        background_tasks.add_task(
            run_generate_job, active_draft_id, config_data, logo_path
        )

        return api_success(
            {
                "ready": False,
                "draft_id": active_draft_id,
                "filename": "",
                "codes": [],
                "warning": "",
                "pdf_url": build_file_url(
                    request, f"/draft/preview/{active_draft_id}", "preview.pdf"
                ),
                "excel_url": None,
            },
            status_code=202,
        )
    except ValidationError as exc:
        LOGGER.warning("Config validation failed: %s", exc)
        shutil.rmtree(draft_dir, ignore_errors=True)
        return api_error(422, "Invalid configuration", details=exc.errors())
    except ApiError as exc:
        LOGGER.warning("Generate request failed: %s", exc.message)
        shutil.rmtree(draft_dir, ignore_errors=True)
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Unexpected error while generating preview")
        shutil.rmtree(draft_dir, ignore_errors=True)
        return api_error(500, "Failed to generate preview")


@app.get("/draft/{draft_id}/status")
def draft_status(draft_id: str, request: Request) -> JSONResponse:
    try:
        draft_dir = get_draft_dir(draft_id)
        if not draft_dir.exists():
            return api_success({"ready": False, "status": "missing"})

        try:
            metadata = read_draft_metadata(draft_id)
            if metadata.get("status") == "error":
                return api_error(
                    500, metadata.get("error") or "Failed to generate preview"
                )
        except FileNotFoundError:
            metadata = {}

        pdf_filename = find_draft_pdf(draft_dir, draft_id)
        export_filename = find_draft_export(draft_dir, draft_id)
        if not pdf_filename:
            return api_success(
                {"ready": False, "status": metadata.get("status") or "processing"}
            )

        pdf_path = safe_child_path(draft_dir, pdf_filename, {".pdf"})
        if not pdf_path.exists():
            return api_success(
                {"ready": False, "status": metadata.get("status") or "processing"}
            )

        return api_success(
            {
                "ready": True,
                "status": "ready",
                "draft_id": draft_id,
                "filename": pdf_filename,
                "warning": metadata.get("warning"),
                **get_draft_asset_urls(
                    request, draft_id, pdf_filename, export_filename
                ),
            }
        )
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to read draft status for %s", draft_id)
        return api_error(500, "Failed to read draft status")


@app.api_route(
    "/draft/preview/{draft_id}/{filename}", methods=["GET", "HEAD"], response_model=None
)
def draft_preview(draft_id: str, filename: str):
    try:
        draft_dir = get_draft_dir(draft_id)
        if not draft_dir.exists():
            return api_error(404, "Draft PDF not found")

        resolved_name = filename
        if Path(filename).name == "preview.pdf":
            resolved_name = find_draft_pdf(draft_dir, draft_id) or ""
        if not resolved_name:
            return api_error(404, "Draft PDF not found")

        file_path = safe_child_path(draft_dir, resolved_name, {".pdf"})
        if not file_path.exists():
            return api_error(404, "Draft PDF not found")

        return FileResponse(file_path, media_type=PDF_MIME_TYPE)
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to serve draft preview %s/%s", draft_id, filename)
        return api_error(500, "Failed to serve draft preview")


@app.api_route(
    "/draft/export/{draft_id}/{filename}", methods=["GET", "HEAD"], response_model=None
)
def draft_export(draft_id: str, filename: str):
    try:
        draft_dir = get_draft_dir(draft_id)
        if not draft_dir.exists():
            return api_error(404, "Draft export not found")

        file_path = safe_child_path(draft_dir, filename, {".xlsx", ".csv"})
        if not file_path.exists():
            return api_error(404, "Draft export not found")

        media_type = (
            CSV_MIME_TYPE if file_path.suffix.lower() == ".csv" else XLSX_MIME_TYPE
        )
        return FileResponse(file_path, media_type=media_type)
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to serve draft export %s/%s", draft_id, filename)
        return api_error(500, "Failed to serve draft export")


@app.post("/draft/confirm/{draft_id}")
def confirm_draft(draft_id: str, request: Request) -> JSONResponse:
    try:
        draft_dir = get_draft_dir(draft_id)
        if not draft_dir.exists():
            raise ApiError(404, "Draft not found")

        metadata = read_draft_metadata(draft_id)
        pdf_filename = str(metadata.get("final_pdf_filename") or "")
        export_filename = str(metadata.get("export_filename") or "") or None
        if not pdf_filename:
            raise ApiError(404, "Draft PDF not found")

        draft_pdf_path = safe_child_path(draft_dir, pdf_filename, {".pdf"})
        if not draft_pdf_path.exists():
            raise ApiError(404, "Draft PDF not found")

        pdf_data = draft_pdf_path.read_bytes()

        export_data = None
        export_mime_type = None
        export_url = None
        if export_filename:
            draft_export_path = safe_child_path(
                draft_dir, export_filename, {".xlsx", ".csv"}
            )
            if draft_export_path.exists():
                export_data = draft_export_path.read_bytes()
                export_mime_type = get_export_mime_type(export_filename)
                export_url = build_file_url(request, "/db/export", export_filename)
            else:
                export_filename = None

        insert_history_record(
            filename=pdf_filename,
            pdf_data=pdf_data,
            export_filename=export_filename,
            export_data=export_data,
            export_mime_type=export_mime_type,
        )

        pdf_url = build_file_url(request, "/db/pdf", pdf_filename)
        shutil.rmtree(draft_dir, ignore_errors=True)
        return api_success(
            {
                "pdf_url": pdf_url,
                "excel_url": export_url,
                "filename": pdf_filename,
                "warning": metadata.get("warning"),
            }
        )
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Unexpected error while confirming draft")
        return api_error(500, "Failed to confirm draft")


@app.delete("/draft/{draft_id}")
def delete_draft(draft_id: str) -> JSONResponse:
    try:
        remove_draft(draft_id)
        return api_success({"message": "Draft deleted"})
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to delete draft %s", draft_id)
        return api_error(500, "Failed to delete draft")


@app.get("/history")
def get_history(request: Request) -> JSONResponse:
    try:
        files = [
            {
                "filename": Path(str(record.get("filename") or "")).name,
                "pdf_url": build_file_url(
                    request,
                    "/db/pdf",
                    Path(str(record.get("filename") or "")).name,
                ),
                "excel_url": (
                    build_file_url(
                        request,
                        "/db/export",
                        Path(str(record.get("export_filename") or "")).name,
                    )
                    if record.get("export_filename")
                    else None
                ),
                "created_at": history_record_timestamp(record.get("created_at")),
            }
            for record in load_history_records()
            if record.get("filename")
        ]

        return api_success({"files": [file for file in files if file["filename"]]})
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to load history")
        return api_error(500, "Failed to load history")


@app.delete("/history/{filename}")
def delete_history_file(filename: str) -> JSONResponse:
    try:
        safe_filename = Path(filename or "").name
        if not safe_filename.lower().endswith(".pdf"):
            raise ApiError(400, "Invalid file type")

        deleted = delete_history_record(safe_filename)
        if not deleted:
            raise ApiError(404, "History file not found")

        return api_success({"message": "Deleted successfully"})
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to delete history file")
        return api_error(500, "Failed to delete history file")


@app.get("/db/pdf/{filename}", response_model=None)
def db_pdf(filename: str):
    try:
        safe_filename = Path(filename or "").name
        if not safe_filename.lower().endswith(".pdf"):
            raise ApiError(400, "Invalid file type")

        pdf_data = load_pdf_data(safe_filename)
        return Response(
            content=pdf_data,
            media_type=PDF_MIME_TYPE,
            headers={
                "Content-Disposition": content_disposition("inline", safe_filename)
            },
        )
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to serve database PDF %s", filename)
        return api_error(500, "Failed to serve PDF")


@app.get("/db/export/{filename}", response_model=None)
def db_export(filename: str):
    try:
        safe_filename = Path(filename or "").name
        if Path(safe_filename).suffix.lower() not in {".xlsx", ".csv"}:
            raise ApiError(400, "Invalid file type")

        export_data, export_mime_type = load_export_data(safe_filename)
        return Response(
            content=export_data,
            media_type=export_mime_type,
            headers={
                "Content-Disposition": content_disposition("attachment", safe_filename)
            },
        )
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to serve database export %s", filename)
        return api_error(500, "Failed to serve export")


@app.api_route("/preview/{filename}", methods=["GET", "HEAD"], response_model=None)
def preview_pdf(filename: str):
    try:
        file_path = safe_child_path(PREVIEW_DIR, filename, {".pdf"})
        if not file_path.exists():
            return api_error(404, "PDF file not found")
        return FileResponse(file_path, media_type=PDF_MIME_TYPE)
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to serve PDF %s", filename)
        return api_error(500, "Failed to serve PDF")


@app.api_route("/excel/{filename}", methods=["GET", "HEAD"], response_model=None)
def preview_excel(filename: str):
    try:
        file_path = safe_child_path(EXCEL_DIR, filename, {".xlsx", ".csv"})
        if not file_path.exists():
            return api_error(404, "Export file not found")
        media_type = (
            CSV_MIME_TYPE if file_path.suffix.lower() == ".csv" else XLSX_MIME_TYPE
        )
        return FileResponse(file_path, media_type=media_type)
    except ApiError as exc:
        return api_error(exc.status_code, exc.message, details=exc.details)
    except Exception:
        LOGGER.exception("Failed to serve export %s", filename)
        return api_error(500, "Failed to serve export")


@app.get("/")
def root(request: Request) -> JSONResponse:
    return api_success(
        {
            "message": "QR Generator API is running",
            "config_page": build_file_url(request, "/page", "config.html"),
        }
    )
